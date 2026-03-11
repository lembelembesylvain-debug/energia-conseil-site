"""
Fusion de 2 fichiers Excel en 1 seul.
- Fichier 1 : DAMIEN_Pipeline_Complet_Mars_2026.xlsx
- Fichier 2 : DAMIEN_Tableau_Prestations_FINAL.xlsx
- Résultat  : DAMIEN_Pipeline_FUSION.xlsx (onglets : Pipeline + Prestations ou tous les onglets des 2 fichiers)
"""
import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
import os

# Chemins des fichiers (même répertoire que le script)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
FILE1 = os.path.join(SCRIPT_DIR, 'DAMIEN_Pipeline_Complet_Mars_2026.xlsx')
FILE2 = os.path.join(SCRIPT_DIR, 'DAMIEN_Tableau_Prestations_FINAL.xlsx')
OUTPUT = os.path.join(SCRIPT_DIR, 'DAMIEN_Pipeline_FUSION.xlsx')

def main():
    if not os.path.isfile(FILE1):
        print("[ERREUR] Fichier introuvable :", FILE1)
        return
    if not os.path.isfile(FILE2):
        print("[ERREUR] Fichier introuvable :", FILE2)
        return

    # Charge les fichiers sources
    wb1 = openpyxl.load_workbook(FILE1)
    wb2 = openpyxl.load_workbook(FILE2)

    # Crée le nouveau fichier
    wb_fusion = openpyxl.Workbook()
    wb_fusion.remove(wb_fusion.active)  # Supprime la feuille par défaut

    # Copie tous les onglets du fichier 1
    for sheet_name in wb1.sheetnames:
        source = wb1[sheet_name]
        target = wb_fusion.create_sheet(title=sheet_name)

        # Copie les cellules
        for row in source:
            for cell in row:
                target[cell.coordinate].value = cell.value
                if cell.has_style:
                    target[cell.coordinate].font = copy(cell.font)
                    target[cell.coordinate].border = copy(cell.border)
                    target[cell.coordinate].fill = copy(cell.fill)
                    target[cell.coordinate].number_format = copy(cell.number_format)
                    target[cell.coordinate].protection = copy(cell.protection)
                    target[cell.coordinate].alignment = copy(cell.alignment)

        # Copie la largeur des colonnes
        for col in source.column_dimensions:
            target.column_dimensions[col].width = source.column_dimensions[col].width

    # Copie tous les onglets du fichier 2
    for sheet_name in wb2.sheetnames:
        source = wb2[sheet_name]
        # Renomme si le nom existe déjà
        new_name = sheet_name
        if new_name in wb_fusion.sheetnames:
            new_name = f"{sheet_name}_2"

        target = wb_fusion.create_sheet(title=new_name)

        # Copie les cellules
        for row in source:
            for cell in row:
                target[cell.coordinate].value = cell.value
                if cell.has_style:
                    target[cell.coordinate].font = copy(cell.font)
                    target[cell.coordinate].border = copy(cell.border)
                    target[cell.coordinate].fill = copy(cell.fill)
                    target[cell.coordinate].number_format = copy(cell.number_format)
                    target[cell.coordinate].protection = copy(cell.protection)
                    target[cell.coordinate].alignment = copy(cell.alignment)

        # Copie la largeur des colonnes
        for col in source.column_dimensions:
            target.column_dimensions[col].width = source.column_dimensions[col].width

    # Enregistre le nouveau fichier
    wb_fusion.save(OUTPUT)

    print("Fusion terminee !")
    print("Fichier cree :", OUTPUT)
    print("Onglets :", ", ".join(wb_fusion.sheetnames))

if __name__ == "__main__":
    main()
