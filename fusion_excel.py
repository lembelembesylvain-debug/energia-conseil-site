"""
Fusion de 2 fichiers Excel en 1 seul.
Sources et sortie : dossier Téléchargements (utilisateur).
"""
import openpyxl
from copy import copy
import os

# Dossier Téléchargements (Windows)
USER_HOME = os.path.expanduser("~")
DOWNLOADS = os.path.join(USER_HOME, "Téléchargements")
if not os.path.isdir(DOWNLOADS):
    DOWNLOADS = os.path.join(USER_HOME, "Downloads")

FILE1 = os.path.join(DOWNLOADS, "DAMIEN_Pipeline_Complet_Mars_2026.xlsx")
FILE2 = os.path.join(DOWNLOADS, "DAMIEN_Tableau_Prestations_FINAL.xlsx")
OUTPUT = os.path.join(DOWNLOADS, "DAMIEN_FUSION.xlsx")


def copy_sheet(source_sheet, target_workbook, sheet_name):
    """Copie une feuille source vers le classeur cible en gardant styles et largeurs."""
    target = target_workbook.create_sheet(title=sheet_name)
    for row in source_sheet:
        for cell in row:
            target_cell = target[cell.coordinate]
            target_cell.value = cell.value
            if cell.has_style:
                target_cell.font = copy(cell.font)
                target_cell.border = copy(cell.border)
                target_cell.fill = copy(cell.fill)
                target_cell.number_format = cell.number_format
                target_cell.protection = copy(cell.protection)
                target_cell.alignment = copy(cell.alignment)
    for col in source_sheet.column_dimensions:
        target.column_dimensions[col].width = source_sheet.column_dimensions[col].width


def main():
    if not os.path.isfile(FILE1):
        print("[ERREUR] Fichier introuvable :", FILE1)
        return
    if not os.path.isfile(FILE2):
        print("[ERREUR] Fichier introuvable :", FILE2)
        return

    wb1 = openpyxl.load_workbook(FILE1)
    wb2 = openpyxl.load_workbook(FILE2)

    wb_fusion = openpyxl.Workbook()
    wb_fusion.remove(wb_fusion.active)

    for name in wb1.sheetnames:
        copy_sheet(wb1[name], wb_fusion, name)

    for name in wb2.sheetnames:
        new_name = name if name not in wb_fusion.sheetnames else f"{name}_2"
        copy_sheet(wb2[name], wb_fusion, new_name)

    wb_fusion.save(OUTPUT)
    print("Fusion terminee. Fichier cree :", OUTPUT)
    print("Onglets :", len(wb_fusion.sheetnames))


if __name__ == "__main__":
    main()
