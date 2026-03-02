#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de génération d'un fichier Excel professionnel pour l'import de 8 clients dans HubSpot CRM.
Commercial rénovation énergétique - Damien
"""

import sys
import io

# Fix encodage console Windows pour les caractères accentués
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# =============================================================================
# DONNÉES DES 8 CLIENTS
# =============================================================================

DEALS_DATA = [
    {
        "nom": "FONTANAY Stéphanie - Rénovation globale",
        "montant_ttc": 64063,
        "profil": "BLEU",
        "aides": 39900,
        "reste_charge": 24163,
        "bilan_mensuel": "+41 €/mois",
        "dpe": "F → C (+3 classes)",
        "date_cloture": "2026-03-27",
        "urgence": "HAUTE",
        "travaux": "ITE 103,8m², isolation plancher 46m², menuiseries, PAC Air/Eau 6kW, ballon thermo",
        "economies_an": 2100,
        "plus_value": 40000,
    },
    {
        "nom": "GUION Quentin & Romane - ITI écologique",
        "montant_ttc": 58000,
        "profil": "ROSE",
        "aides": 20400,
        "reste_charge": 37600,
        "bilan_mensuel": "+41 €/mois",
        "dpe": "F → B (+4 classes)",
        "date_cloture": "2026-03-24",
        "urgence": "MOYENNE",
        "travaux": "ITI fibre bois/ouate 100m², plancher 80m², menuiseries, VMC double flux, PAC Air/Eau 8-10kW, ballon thermo",
        "economies_an": 3000,
        "plus_value": 40000,
    },
    {
        "nom": "ROYER Maixent - Rénovation toiture",
        "montant_ttc": 52000,
        "profil": "JAUNE",
        "aides": 25550,
        "reste_charge": 26450,
        "bilan_mensuel": "N/A",
        "dpe": "G → D (+3 classes)",
        "date_cloture": "2026-04-30",
        "urgence": "TRÈS HAUTE",
        "travaux": "Isolation toiture 80m² + couverture, PAC Air-Air 3 splits Daikin, ballon thermo 200L, 5 fenêtres + 2 portes, zinguerie",
        "economies_an": 0,
        "plus_value": 0,
    },
    {
        "nom": "YANG Simiao & LIU Ke - Rénovation globale",
        "montant_ttc": 120743,
        "profil": "BLEU",
        "aides": 67500,
        "reste_charge": 53243,
        "bilan_mensuel": "-121 €/mois (2 ans) puis +17 €/mois",
        "dpe": "F/G → B (+5 classes)",
        "date_cloture": "2026-03-24",
        "urgence": "HAUTE",
        "travaux": "ITE 332m², rampants 138m², plancher 75m², PAC Air/Eau 17kW, VMC hygro B, ballon 240L",
        "economies_an": 2650,
        "plus_value": 60000,
    },
    {
        "nom": "SEUX Laurent - Rénovation globale",
        "montant_ttc": 70391,
        "profil": "JAUNE",
        "aides": 39000,
        "reste_charge": 31391,
        "bilan_mensuel": "+6 €/mois",
        "dpe": "F → B (+4 classes)",
        "date_cloture": "2026-03-27",
        "urgence": "HAUTE",
        "travaux": "ITE 78,25m², combles 56m², 6 fenêtres + 2 portes-fenêtres + 1 porte, PAC Air/Eau, VMC hygro B, désembouage, électricité",
        "economies_an": 1950,
        "plus_value": 25000,
    },
    {
        "nom": "BRUNEL Emrick - Rénovation ITE (Indivision)",
        "montant_ttc": 71972,
        "profil": "JAUNE",
        "aides": 39500,
        "reste_charge": 32472,
        "bilan_mensuel": "-256 €/mois",
        "dpe": "F → D (+2 classes)",
        "date_cloture": "2026-04-25",
        "urgence": "MOYENNE",
        "travaux": "ITE 196m², combles 72m², plancher 48m², 1 porte-fenêtre + 4 fenêtres + volets solaires, VMC hygro, 5 splits PAC air/air, ballon thermo 200L",
        "economies_an": 1800,
        "plus_value": 30000,
    },
    {
        "nom": "GRANGE Mathias - Rénovation ITI pisé",
        "montant_ttc": 52000,
        "profil": "JAUNE",
        "aides": 27750,
        "reste_charge": 24250,
        "bilan_mensuel": "-27 €/mois (15 ans) puis +108 €/mois",
        "dpe": "G → D/C (+3/4 classes)",
        "date_cloture": "2026-04-30",
        "urgence": "FAIBLE",
        "travaux": "Isolation combles 60m², ITI laine roche 100m², 5 fenêtres + 5 volets solaires + porte, PAC Air-Air 5 splits, VMC hygro, ballon thermo 200L",
        "economies_an": 1300,
        "plus_value": 0,
    },
    {
        "nom": "RICHARD Paul - Reprise KAZEO",
        "montant_ttc": 56000,
        "profil": "ROSE",
        "aides": 28606,
        "reste_charge": 27394,
        "bilan_mensuel": "272 €/mois (financement)",
        "dpe": "G → A (+6 classes)",
        "date_cloture": "2026-08-21",
        "urgence": "CRITIQUE",
        "travaux": "ITE 164m², combles 100m², PAC Air/Air 4 splits, ballon thermo 200L, VMC hygro B",
        "economies_an": 2400,
        "plus_value": 80000,
    },
]

CONTACTS_DATA = [
    {"prenom": "Stéphanie", "nom": "FONTANAY", "email": "fontanay-stephanie@orange.fr", "tel": "06 59 38 97 49",
     "adresse": "167 Rue des Acacias", "ville": "MONTROND-LES-BAINS", "cp": "42210", "profil": "BLEU", "projet": "Rénovation globale"},
    {"prenom": "Quentin & Romane", "nom": "GUION", "email": "quentin-guion@hotmail.fr", "tel": "",
     "adresse": "", "ville": "", "cp": "", "profil": "ROSE", "projet": "ITI écologique"},
    {"prenom": "Maixent", "nom": "ROYER", "email": "À compléter", "tel": "À compléter",
     "adresse": "6 route de Marcilly", "ville": "CHALAIN D'UZOR", "cp": "42600", "profil": "JAUNE", "projet": "Rénovation toiture"},
    {"prenom": "Simiao & Ke", "nom": "YANG-LIU", "email": "yespace2016@gmail.com", "tel": "07 64 42 85 25",
     "adresse": "18 Parc des Chavannes", "ville": "COLLONGES-AU-MONT-D'OR", "cp": "69660", "profil": "BLEU", "projet": "Rénovation globale"},
    {"prenom": "Laurent", "nom": "SEUX", "email": "", "tel": "",
     "adresse": "4 Rue Georges Laurent", "ville": "LA RICAMARIE", "cp": "42150", "profil": "JAUNE", "projet": "Rénovation globale"},
    {"prenom": "Emrick", "nom": "BRUNEL", "email": "", "tel": "",
     "adresse": "48 Impasse des Lagunes", "ville": "BUSSIÈRES", "cp": "42510", "profil": "JAUNE", "projet": "Rénovation ITE (Indivision)"},
    {"prenom": "Mathias", "nom": "GRANGE", "email": "", "tel": "06 67 52 88 90",
     "adresse": "569 Chemin des Places", "ville": "MONTCHAL", "cp": "42360", "profil": "JAUNE", "projet": "Rénovation ITI pisé"},
    {"prenom": "Paul", "nom": "RICHARD", "email": "", "tel": "",
     "adresse": "600 Route de la Goutte", "ville": "UNIAS", "cp": "42210", "profil": "ROSE", "projet": "Reprise KAZEO"},
]

# =============================================================================
# STYLES
# =============================================================================

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

HEADER_DEALS = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Bleu foncé
HEADER_CONTACTS = PatternFill(start_color="2E5C3E", end_color="2E5C3E", fill_type="solid")  # Vert foncé
HEADER_SYNTHESE = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")  # Orange

URGENCE_COLORS = {
    "CRITIQUE": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),  # Rouge
    "TRÈS HAUTE": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),  # Rouge
    "HAUTE": PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid"),  # Orange
    "MOYENNE": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),  # Jaune
    "FAIBLE": PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),  # Vert
}


def format_euro(value):
    """Formate un nombre en euros avec séparateurs de milliers."""
    if value == 0 or value == "":
        return "0 €" if value == 0 else ""
    return f"{int(value):,} €".replace(",", " ")


def apply_borders(ws, max_row, max_col):
    """Applique les bordures à une plage de cellules."""
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER


def create_deals_sheet(wb):
    """Crée l'onglet Deals avec formatage professionnel."""
    ws = wb.active
    ws.title = "Deals"

    headers = [
        "Nom du deal", "Montant TTC", "Profil", "Aides", "Reste à charge", "Bilan mensuel",
        "DPE", "Date clôture", "Urgence", "Travaux", "Économies/an", "Plus-value"
    ]

    # Headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_DEALS
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Données
    for row_idx, deal in enumerate(DEALS_DATA, 2):
        ws.cell(row=row_idx, column=1, value=deal["nom"]).alignment = Alignment(horizontal="left", wrap_text=True)
        ws.cell(row=row_idx, column=2, value=format_euro(deal["montant_ttc"])).alignment = Alignment(horizontal="right")
        ws.cell(row=row_idx, column=3, value=deal["profil"]).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=4, value=format_euro(deal["aides"])).alignment = Alignment(horizontal="right")
        ws.cell(row=row_idx, column=5, value=format_euro(deal["reste_charge"])).alignment = Alignment(horizontal="right")
        ws.cell(row=row_idx, column=6, value=deal["bilan_mensuel"]).alignment = Alignment(horizontal="left")
        ws.cell(row=row_idx, column=7, value=deal["dpe"]).alignment = Alignment(horizontal="left")
        ws.cell(row=row_idx, column=8, value=deal["date_cloture"]).alignment = Alignment(horizontal="center")
        cell_urgence = ws.cell(row=row_idx, column=9, value=deal["urgence"])
        cell_urgence.alignment = Alignment(horizontal="center")
        cell_urgence.fill = URGENCE_COLORS.get(deal["urgence"], PatternFill())
        ws.cell(row=row_idx, column=10, value=deal["travaux"]).alignment = Alignment(horizontal="left", wrap_text=True)
        ws.cell(row=row_idx, column=11, value=format_euro(deal["economies_an"])).alignment = Alignment(horizontal="right")
        ws.cell(row=row_idx, column=12, value=format_euro(deal["plus_value"])).alignment = Alignment(horizontal="right")

    # Ligne de total
    total_row = len(DEALS_DATA) + 2
    ws.cell(row=total_row, column=1, value="TOTAL PIPELINE").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=format_euro(sum(d["montant_ttc"] for d in DEALS_DATA))).font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=format_euro(sum(d["aides"] for d in DEALS_DATA))).font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=format_euro(sum(d["reste_charge"] for d in DEALS_DATA))).font = Font(bold=True)
    ws.cell(row=total_row, column=11, value=format_euro(sum(d["economies_an"] for d in DEALS_DATA))).font = Font(bold=True)
    ws.cell(row=total_row, column=12, value=format_euro(sum(d["plus_value"] for d in DEALS_DATA))).font = Font(bold=True)
    for col in range(1, 13):
        ws.cell(row=total_row, column=col).alignment = Alignment(horizontal="right" if col in [2, 4, 5, 11, 12] else "left")
        ws.cell(row=total_row, column=col).fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")

    apply_borders(ws, total_row, 12)

    # Ajustement largeur colonnes
    column_widths = [35, 14, 10, 14, 14, 25, 18, 14, 12, 45, 14, 14]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_contacts_sheet(wb):
    """Crée l'onglet Contacts."""
    ws = wb.create_sheet("Contacts", 1)

    headers = ["Prénom", "Nom", "Email", "Téléphone", "Adresse", "Ville", "Code postal", "Profil", "Projet"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_CONTACTS
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    for row_idx, contact in enumerate(CONTACTS_DATA, 2):
        ws.cell(row=row_idx, column=1, value=contact["prenom"]).alignment = Alignment(horizontal="left")
        ws.cell(row=row_idx, column=2, value=contact["nom"]).alignment = Alignment(horizontal="left")
        ws.cell(row=row_idx, column=3, value=contact["email"]).alignment = Alignment(horizontal="left")
        ws.cell(row=row_idx, column=4, value=contact["tel"]).alignment = Alignment(horizontal="left")
        ws.cell(row=row_idx, column=5, value=contact["adresse"]).alignment = Alignment(horizontal="left")
        ws.cell(row=row_idx, column=6, value=contact["ville"]).alignment = Alignment(horizontal="left")
        ws.cell(row=row_idx, column=7, value=contact["cp"]).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=8, value=contact["profil"]).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=9, value=contact["projet"]).alignment = Alignment(horizontal="left")

    apply_borders(ws, len(CONTACTS_DATA) + 1, 9)

    column_widths = [18, 15, 28, 16, 30, 25, 12, 10, 25]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_synthese_sheet(wb):
    """Crée l'onglet Synthèse avec KPIs et graphique."""
    ws = wb.create_sheet("Synthèse", 2)

    total_ttc = sum(d["montant_ttc"] for d in DEALS_DATA)
    total_aides = sum(d["aides"] for d in DEALS_DATA)
    total_reste = sum(d["reste_charge"] for d in DEALS_DATA)
    taux_prise_charge = (total_aides / total_ttc * 100) if total_ttc > 0 else 0

    # Profils
    profils = {}
    for d in DEALS_DATA:
        profils[d["profil"]] = profils.get(d["profil"], 0) + 1

    # Urgences
    urgences = {}
    for d in DEALS_DATA:
        urgences[d["urgence"]] = urgences.get(d["urgence"], 0) + 1

    # Clients prioritaires
    prioritaires = [d["nom"] for d in DEALS_DATA if d["urgence"] in ("TRÈS HAUTE", "HAUTE", "CRITIQUE")]

    row = 1

    # Titre
    ws.cell(row=row, column=1, value="SYNTHÈSE PIPELINE - 8 CLIENTS HUBSPOT").font = Font(bold=True, size=14)
    row += 2

    # KPIs
    kpis = [
        ("Total pipeline (TTC)", format_euro(total_ttc)),
        ("Total aides", format_euro(total_aides)),
        ("Total reste à charge", format_euro(total_reste)),
        ("Taux moyen de prise en charge", f"{taux_prise_charge:.1f} %"),
    ]
    for label, value in kpis:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws.cell(row=row, column=2).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        row += 1

    row += 1

    # Répartition par profil
    ws.cell(row=row, column=1, value="Répartition par profil MaPrimeRénov'").font = Font(bold=True)
    row += 1
    for profil, count in sorted(profils.items()):
        ws.cell(row=row, column=1, value=f"  {profil}")
        ws.cell(row=row, column=2, value=f"{count} client(s)")
        row += 1

    row += 1

    # Répartition par urgence (données pour graphique)
    ws.cell(row=row, column=1, value="Répartition par urgence").font = Font(bold=True)
    row += 1
    urgence_order = ["CRITIQUE", "TRÈS HAUTE", "HAUTE", "MOYENNE", "FAIBLE"]
    for urg in urgence_order:
        if urg in urgences:
            ws.cell(row=row, column=1, value=urg)
            ws.cell(row=row, column=2, value=urgences[urg])
            row += 1

    chart_start_row = row - len([u for u in urgence_order if u in urgences])
    chart_end_row = row - 1

    row += 2

    # Graphique en barres
    if chart_end_row >= chart_start_row:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Répartition par niveau d'urgence"
        chart.y_axis.title = "Nombre de clients"
        chart.x_axis.title = "Urgence"
        data = Reference(ws, min_col=2, min_row=chart_start_row, max_row=chart_end_row)
        cats = Reference(ws, min_col=1, min_row=chart_start_row, max_row=chart_end_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 12
        chart.height = 8
        ws.add_chart(chart, f"D{chart_start_row}")

    row = chart_end_row + 15 if chart_end_row >= chart_start_row else row + 5

    # Actions prioritaires
    ws.cell(row=row, column=1, value="Actions prioritaires (urgence HAUTE et plus)").font = Font(bold=True, color="C00000")
    row += 1
    for i, client in enumerate(prioritaires, 1):
        ws.cell(row=row, column=1, value=f"  {i}. {client}")
        row += 1

    apply_borders(ws, row, 2)

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 20


def main():
    """Point d'entrée principal."""
    print("Génération du fichier Excel HubSpot en cours...")

    wb = Workbook()

    create_deals_sheet(wb)
    create_contacts_sheet(wb)
    create_synthese_sheet(wb)

    filename = "Damien_8_Clients_HubSpot_Import.xlsx"
    wb.save(filename)

    print(f"\n[OK] Fichier généré avec succès : {filename}")
    print("   - Onglet 'Deals' : 8 deals avec codes couleur urgence")
    print("   - Onglet 'Contacts' : 8 contacts prêts pour l'import")
    print("   - Onglet 'Synthèse' : KPIs, répartitions et actions prioritaires")
    print("\nLe fichier est prêt pour l'import dans HubSpot CRM.")


if __name__ == "__main__":
    main()
